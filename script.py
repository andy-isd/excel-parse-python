import os
import openpyxl
import xlrd
from openpyxl import load_workbook, Workbook

def extract_target_row_xlsx(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    target_row = None
    found_rajion = False
    index_rajion = None

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row[1] == 'Район' or row[1] == 'Територіальний підрозділ':
            found_rajion = True
            index_rajion = index
        elif found_rajion and index == index_rajion + 4:
            if row[1] == 2:
                index_rajion = index_rajion + 1
            else:
                target_row = list(row)

    workbook.close()

    return target_row

def extract_target_row_xls(file_path):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)

    target_row = None
    found_rajion = False
    index_rajion = None

    for row_index in range(sheet.nrows):
        row = sheet.row_values(row_index)
        if row[1] == 'Район' or row[1] == 'Територіальний підрозділ':
            found_rajion = True
            index_rajion = row_index
        elif found_rajion and row_index == index_rajion + 4:
            if row[1] == 2:
                index_rajion = index_rajion + 1
            else:
                target_row = row
            break

    return target_row

def extract_target_row(file_path):
    if file_path.endswith('.xlsx'):
        return extract_target_row_xlsx(file_path)
    elif file_path.endswith('.xls'):
        return extract_target_row_xls(file_path)
    else:
        print(f"Unsupported file format: {file_path}")
        return None

def find_excel_files(input_folder):
    excel_files = []
    for root, _, files in os.walk(input_folder):
        for filename in files:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                file_path = os.path.join(root, filename)
                excel_files.append(file_path)
    return excel_files

def main():
    input_folder = 'input/'  # Change the default input folder name to 'input/'
    output_file = 'output.xlsx'

    target_rows = []
    total_files_processed = 0

    excel_files = find_excel_files(input_folder)

    for file_path in excel_files:
        print(f"Opening {file_path}...")
        target_row = extract_target_row(file_path)
        if target_row:
            target_rows.append(target_row)
            total_files_processed += 1
            print(f"OK")
        else:
            print(f"Target row not found")

    if target_rows:
        workbook = Workbook()
        sheet = workbook.active

        for row in target_rows:
            sheet.append(row)

        workbook.save(output_file)
        print(f"Extracted {len(target_rows)} target rows and saved to {output_file}")
    else:
        print("No target rows found in any files.")

    print(f"Total number of files processed: {total_files_processed}")

if __name__ == '__main__':
    main()
